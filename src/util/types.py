from __future__ import annotations  # 無くていいはずなんだが python 3.12.8
from enum import Enum
import inspect
from importlib.util import spec_from_file_location, module_from_spec
from pathlib import Path
from datetime import datetime
from collections import namedtuple
from dataclasses import dataclass
import sys
from typing import Iterator
from pprint import pprint
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ResultCount = namedtuple("ResultCount", ["succeeded", "failed"])


class TestScope(Enum):
    ALL = "all"  # all testcases
    LAST_FAILED = "lastfailed"  # only failed testcases in last run


@dataclass
class TestCase:
    """個々のテストケース
    testid: テストモジュールの識別ID
    group: テストグループの名称
    module: テストを実装したモジュール
    testfunction: テストケースとして実行する関数名
    subject: テストケースについての説明。任意
    ignore: このテストケースを無視するかどうか
    """

    testid: str
    group: str
    module: TestModule
    testfunction: str
    subject: str
    ignore: bool


@dataclass
class TestResult:
    """テストケースの実行結果
    testid: テストモジュールの識別ID
    group: テストグループの名称
    module: テストを実装したモジュールの相対パス
    testfunction: テストケースとして実行する関数名
    succeeded: テストケースが成功したかどうか
    runned_at: テストケースが実行された日時
    """

    testid: str
    group: str
    module: str
    testfunction: str
    succeeded: bool
    runned_at: str


class TestModule:
    """テストを実装したモジュール"""

    testid: str

    def __init__(
        self, testid: str, subject: str, group: str, modulepath: str, run: bool
    ) -> None:
        """テストモジュールを作成する。
        testid: テストモジュールの識別ID
        subject: テストモジュールの説明
        group: テストグループの名称
        modulepath: テストモジュールのパス
        run: テストモジュールを実行するかどうか
        """
        self.testid = testid
        self.__subject = subject
        self.__group = group
        self.__modulepath = modulepath
        self.__run = run

        self.__testcases: list[TestCase] = []
        self.__testfunctions: list[str] = []
        self.__results: dict[str, TestResult] = {}

        self.__testmodule = None
        self.__testmodulekey = "testee"

    def load_module(self, key=datetime.now()) -> None:
        if self.__testmodulekey != "testee":
            self.unload_module()
        if not Path(self.modulepath).exists():
            return  # ロード失敗時はスキップ
        spec = spec_from_file_location("testee", str(self.modulepath))
        mod = module_from_spec(spec)
        spec.loader.exec_module(mod)
        self.__testmodule = mod
        self.__testmodulekey = "testee" + str(key)
        sys.modules[self.__testmodulekey] = mod

    @property
    def testmodule(self):
        """ロードされたテストモジュールを返す"""
        return self.__testmodule

    @property
    def group(self) -> str:
        return self.__group

    @property
    def subject(self):
        return self.__subject

    @property
    def modulepath(self) -> Path:
        return Path(self.__modulepath).resolve()

    @property
    def run(self) -> bool:
        return self.__run

    def unload_module(self) -> None:
        if self.__testmodule:
            del self.__testmodule
            self.__testmodule = None
        if self.__testmodulekey in sys.modules:
            del sys.modules[self.__testmodulekey]
        self.__testmodulekey = "testee"

    def add_testcase(self, testfunction: str, subject: str, ignore: bool) -> TestCase:
        tc = TestCase(
            testid=self.testid,
            group=self.__group,
            module=self.__modulepath,
            testfunction=testfunction,
            subject=subject,
            ignore=ignore,
        )
        self.__testcases.append(tc)
        self.__testfunctions.append(testfunction)
        return tc

    def pick_testcases(self) -> None:
        """テストモジュール内の全てのテストケースを列挙する"""
        self.load_module()
        testfunctions = [
            (name, f)
            for name, f in inspect.getmembers(self.__testmodule, inspect.isfunction)
            if name.startswith("test_")
        ]
        for name, f in testfunctions:
            self.add_testcase(
                testfunction=name,
                subject=getattr(f, "_subject", ""),
                ignore=getattr(f, "_is_ignored", False),
            )
        self.unload_module()

    def set_result(
        self, testfunction: str, succeeded: bool, runned_at: datetime
    ) -> TestResult | None:
        if testfunction in self.__testfunctions:
            succeeded = TestResult(
                testid=self.testid,
                group=self.__group,
                module=self.__modulepath,
                testfunction=testfunction,
                succeeded=succeeded,
                runned_at=runned_at,
            )
            self.__results[self.__get_result_key(self.testid, testfunction)] = succeeded
            return succeeded
        else:
            return None

    def get_result(self, testfunction: str) -> TestResult | None:
        if self.__get_result_key(self.testid, testfunction) in self.__results:
            return self.__results[self.__get_result_key(self.testid, testfunction)]
        else:
            return None

    def __get_result_key(self, testid: str, testfunction: str) -> str:
        return testfunction + "@" + testid

    @property
    def results(self) -> list[TestResult]:
        sortedresults = [self.__results[k] for k in sorted(self.__results)]
        return sortedresults

    @property
    def resultcount(self) -> ResultCount:
        """成功と失敗の数を返す"""
        success = 0
        fail = 0
        for result in self.__results.values():
            if result.succeeded:
                success += 1
            else:
                fail += 1
        return ResultCount(succeeded=success, failed=fail)

    def __iter__(self) -> Iterator[TestCase]:
        """テストモジュールに含まれるテストケースのリスト"""
        return iter(self.__testcases)

    def __getitem__(self, index: int | str) -> TestCase | None:
        if isinstance(index, int):
            if index < 0 or len(self.__testcases) <= index:
                return None
            return self.__testcases[index]
        else:
            for testcase in self.__testcases:
                if testcase.testfunction == index:
                    return testcase
            return None

    @property
    def count(self) -> int:
        return len(self.__testcases)


class TestGroup:
    """意味的にまとまったテストケースのグループ。シナリオファイルの1つのシートに対応する
    groupname: グループ名
    """

    groupname: str

    def __init__(self, name: str) -> None:
        """グループの初期化
        name: グループ名。シナリオファイルのシート名
        """
        self.groupname = name
        self.__testmodules: list[TestModule] = []

    def add_test_module(
        self, testid: str, module: str, subject: str, run: bool
    ) -> TestModule:
        testmodule = TestModule(
            testid=testid,
            subject=subject,
            group=self.groupname,
            modulepath=module,
            run=run,
        )
        self.__testmodules.append(testmodule)
        return testmodule

    def __iter__(self) -> Iterator[TestModule]:
        """グループにかかわらず全ての定義されたテストモジュールのリスト"""
        return iter(self.__testmodules)

    def __getitem__(self, index: int | str) -> TestModule | None:
        if isinstance(index, int):
            if index < 0 or len(self.__testmodules) <= index:
                return None
            return self.__testmodules[index]
        else:
            for module in self.__testmodules:
                if module.testid == index:
                    return module
            return None

    @property
    def count(self) -> int:
        return len(self.__testmodules)


class TestScenario:
    """テストシナリオの定義。シナリオファイルに対応する"""

    def __init__(self, scenariopath: Path) -> None:
        """テストシナリオの初期化
        scenariopath: シナリオファイルの絶対パス
        """
        if not scenariopath.exists():
            print(f"[ERR]file not exists. {str(scenariopath)}")
            self.__valid = False
            return

        self.__valid = True
        self.__scenario = scenariopath
        self.__groups: list[TestGroup] = []
        self.__groupsindex: dict[str, TestGroup] = {}

        # ファイル読み込み
        sbook = None
        try:
            print(f"open scenario file: {self.__scenario}")
            sbook = load_workbook(self.__scenario, read_only=True)

            for gsheet in sbook.worksheets:
                group = self.__analyzegroup(gsheet=gsheet)
                if group is None:
                    print(f"[WARN]sheet '{gsheet.title}' is not valid test definition.")
                    continue
                self.__groups.append(group)
                self.__groupsindex[group.groupname] = group

            if len(self.__groups) == 0:
                print(f"[ERR]No test found in '{self.__scenario}'.")
                self.__valid = False

        except ValueError as ve:
            print("Scenario parse value error :")
            pprint(ve)
            self.__valid = False
        except Exception as e:
            print("Scenario parse unknown error :")
            pprint(e)
            self.__valid = False
        finally:
            if sbook:
                print("close scenario file.")
                sbook.close()

    def __iter__(self) -> Iterator[TestGroup]:
        return iter(self.__groups)

    def __getitem__(self, index: int) -> TestGroup | None:
        if index < 0 or len(self.__groups) <= index:
            return None
        return self.__groups[index]

    @property
    def valid(self) -> bool:
        return self.__valid

    @property
    def count(self) -> int:
        return len(self.__groups)

    @property
    def path(self) -> Path:
        return self.__scenario

    def group(self, groupname: str) -> TestGroup:
        if groupname not in self.__groupsindex:
            raise ValueError(f"[ERR]Unknown group name '{groupname}'.")

        return self.__groupsindex[groupname]

    def __analyzegroup(self, gsheet: Worksheet) -> TestGroup | None:
        if gsheet.cell(2, 2).value != "テストID":
            return None
        if gsheet.cell(2, 3).value is None or gsheet.cell(2, 3).value != "説明":
            return None
        if gsheet.cell(2, 4).value is None or gsheet.cell(2, 4).value != "モジュール":
            return None
        if gsheet.cell(2, 5).value is None or gsheet.cell(2, 5).value != "実行":
            return None
        if gsheet.cell(2, 6).value is None or gsheet.cell(2, 6).value != "結果":
            return None

        group = TestGroup(gsheet.title)
        row = 3
        while gsheet.cell(row, 2).value is not None and gsheet.cell(row, 2).value != "":
            group.add_test_module(
                testid=gsheet.cell(row, 2).value,
                subject=gsheet.cell(row, 3).value,
                module=gsheet.cell(row, 4).value,
                run=self.__getrequired(gsheet.cell(row, 5).value),
            )
            row += 1

        return group

    def __getrequired(self, runvalue: str) -> bool:
        return runvalue == "○"


@dataclass
class TestSuite:
    """まとめて実行するテストケースのまとまり
    name: テストスイートの名前。ログに書かれるのと、「前回の実行」を識別するもの
    subject: テストスイートの説明。任意
    scope: テストスイートの実行範囲。デフォルトはFULL
    tests: 実行するテストセットのリスト
    """

    def __init__(
        self,
        name: str,
        subject: str,
        scenario: TestScenario,
        filters: str = "",
        ignores: str = "",
        scope: TestScope = TestScope.ALL,
    ):
        """テストセットを作成する。
        name: テストスイートを識別する名称（この名前でログを出力する）
        subject: テストスイートの説明
        scenario: テストシナリオ
        filters: 実行するべきテストケースの関数名に含まれる識別子（任意の文字列）。"|"区切り。省略した場合はフィルタリングしない
        ignores: 無視するテストケースの関数名に含まれる識別子（任意の文字列）。"|"区切り。省略した場合は無視しない
        scope: 実行する範囲。全部か前回の失敗のみ
        """
        self.__name = name
        self.__subject = subject
        self.__filters = self.__getconditionlist(filters)
        self.__ignores = self.__getconditionlist(ignores)
        self.__scope = scope
        self.__tests = list[TestCase]()
        self.__testset = dict[str, list[TestCase]]()
        for group in scenario:
            for module in group:
                if not module.run:
                    continue
                self.__testset[module.testid] = list[TestCase]()
                module.pick_testcases()
                for testcase in module:
                    if self.__torun(
                        testcase=testcase,
                        filters=self.__filters,
                        ignores=self.__ignores,
                    ):
                        self.__tests.append(testcase)
                        self.__testset[module.testid].append(testcase)

    @property
    def name(self) -> str:
        return self.__name

    @property
    def subject(self) -> str:
        return self.__subject

    @property
    def allscope(self) -> bool:
        return self.__scope == TestScope.ALL

    @property
    def lastfailedscope(self) -> bool:
        return self.__scope == TestScope.LAST_FAILED

    @property
    def count(self) -> int:
        return len(self.__tests)

    def __getconditionlist(self, conditions: str) -> list[str]:
        return [f.strip() for f in conditions.split("|") if len(f.strip()) > 0]

    def __torun(
        self, testcase: TestCase, filters: list[str], ignores: list[str]
    ) -> bool:
        if testcase.ignore:
            return False
        if self.__includestest(testcase.testfunction, ignores):
            return False
        if len(filters) == 0 or self.__includestest(testcase.testfunction, filters):
            return True

        return False

    def __includestest(self, testname: str, condition: list[str]) -> bool:
        if not condition or len(condition) == 0:
            return False
        for f in condition:
            if f in testname:
                return True
        return False

    def __iter__(self) -> Iterator[TestCase]:
        return iter(self.__tests)

    def __getitem__(self, index: int | str) -> list[TestCase] | None:
        if isinstance(index, str):
            # testidからテストケースのリストを取出す
            if index in self.__testset:
                return self.__testset[index]
            return None
        else:
            # indexの位置のテストケース単体を含むリストを返す
            if index < 0 or len(self.__tests) <= index:
                return None
            return [self.__tests[index]]
