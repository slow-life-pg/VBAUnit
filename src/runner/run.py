from pathlib import Path
import shutil
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from natsort import natsort_keygen
from util.types import TestSuite, TestModule, TestCase, TestResult


def __gettimestampnow() -> str:
    return datetime.now().isoformat(sep=" ", timespec="milliesconds")


def __writestartlog(testcase: TestCase, f) -> None:
    config: dict[str, str] = {
        "testid": testcase.testid,
        "group": testcase.group,
        "module": testcase.module.modulepath.name,
        "function": testcase.testfunction,
        "at": str(testcase.start_line),
        "start": __gettimestampnow(),
    }
    f.write(json.dumps(config) + "\n")
    print("Running test case:")
    print(json.dumps(config))


def __writeendlog(result: TestResult, f) -> None:
    outcome: dict[str, str] = {
        "outcome": str(result.succeeded),
        "runned": result.runned_at,
    }
    f.write(json.dumps(outcome) + "\n")
    print(json.dumps(outcome))


def __createresult(testcase: TestCase, succeeded: bool) -> TestResult:
    result = TestResult(
        testid=testcase.testid,
        group=testcase.group,
        module=testcase.module,
        testfunction=testcase.testfunction,
        start_line=testcase.start_line,
        succeeded=succeeded,
        runned_at=__gettimestampnow(),
    )
    return result


def __runtestcase(testcase: TestCase, f) -> TestResult:
    return __createresult(testcase=testcase, succeeded=False)


def __setmoduleresults(
    testcase: TestCase,
    result: TestResult,
    modulelist: list[TestModule],
    modulesummary_success: dict[str, int],
    modulesummary_failure: dict[str, int],
) -> None:
    modulename = testcase.module.modulepath.name
    if modulename not in modulesummary_success:
        modulesummary_success[modulename] = 0
        modulesummary_failure[modulename] = 0
        modulelist.append(testcase.module)
    if result.succeeded:
        modulesummary_success[modulename] += 1
    else:
        modulesummary_failure[modulename] += 1


def __runtestsuite(
    suite: TestSuite,
    results: list[TestResult],
    modulelist: list[TestModule],
    f,
    modulesummary_success: dict[str, int],
    modulesummary_failure: dict[str, int],
) -> None:
    for testcase in suite:
        __writestartlog(testcase=testcase, f=f)

        try:
            result = __runtestcase(testcase=testcase, f=f)
        except Exception as e:
            print(f"Runtime error: {e}")
            result = __createresult(testcase=testcase, succeeded=False)

        __writeendlog(result=result, f=f)
        f.flush()

        results.append(result)
        __setmoduleresults(
            testcase=testcase,
            result=result,
            modulelist=modulelist,
            modulesummary_success=modulesummary_success,
            modulesummary_failure=modulesummary_failure,
        )


def __writemodulersult(
    resultbook: Workbook, modulelist: list[TestModule], modulesummary_success: dict[str, int], modulesummary_failure: dict[str, int]
) -> None:
    for module in modulelist:
        modulename = module.modulepath.name
        resultsheet = resultbook[module.group]
        if modulesummary_success[modulename] >= 0 and modulesummary_failure[modulename] == 0:
            resultsheet.cell(row=module.line, column=6, value="○")
        elif modulesummary_success[modulename] > 0 and modulesummary_failure[modulename] > 0:
            resultsheet.cell(row=module.line, column=6, value="△")
        else:
            resultsheet.cell(row=module.line, column=6, value="×")


def __writeresults(resultsheet: Worksheet, results: list[TestResult]) -> None:
    max_row = 1
    rkey = natsort_keygen()
    for result in sorted(results, key=lambda r: (rkey(r.testid), r.testfunction)):
        resultdump = {
            "testid": result.testid,
            "module": result.module.modulepath.name,
            "function": result.testfunction,
            "line": result.start_line,
            "succeeded": result.succeeded,
            "runned_at": result.runned_at,
        }
        newsheet.cell(row=max_row, column=1, value=json.dumps(resultdump))
        max_row += 1


def run_testsuite(suite: TestSuite, scenario: Path, bridge: Path, out: Path) -> None:
    outputpath = out.joinpath(scenario.name)
    testlogpath = out.joinpath("testlog.txt")

    with open(testlogpath, mode="w", encoding="utf-8") as f:
        f.write(f"Test log for {suite.name}\n")
        f.write(f"Subject: {suite.subject}\n")
        f.write(f"Scenario: {scenario}\n")
        f.write(f"Output: {out}\n")
        f.write(f"Bridge: {bridge}\n")
        f.write(f"Started at: {__gettimestampnow()}\n")
        f.write("\n")
    print(f"Running test suite for scenario: {suite.name}")
    print(f"{suite.subject}")
    print(f"{scenario}")
    print(f"Output: {out}")

    shutil.copy(scenario, outputpath)

    # 実行

    results = list[TestResult]()
    modulelist = list[TestModule]()
    modulesummary_success: dict[str, int] = {}
    modulesummary_failure: dict[str, int] = {}
    with open(testlogpath, mode="a", encoding="utf-8") as f:
        f.write(f"\nStart at: {__gettimestampnow()}\n")
        __runtestsuite(
            suite=suite,
            results=results,
            modulelist=modulelist,
            f=f,
            modulesummary_success=modulesummary_success,
            modulesummary_failure=modulesummary_failure,
        )
        f.write(f"\nFinished at: {__gettimestampnow()}\n")

    # 結果の書込

    resultbook = None
    try:
        resultbook = load_workbook(outputpath)

        # モジュールの結果
        __writemodulersult(
            resultbook=resultbook,
            modulelist=modulelist,
            modulesummary_success=modulesummary_success,
            modulesummary_failure=modulesummary_failure,
        )

        # テストケースの結果
        newsheetname = f"Results_{datetime.now():%Y%m%d_%H%M%S}"
        newsheet = resultbook.create_sheet(newsheetname)
        __writeresults(resultsheet=newsheet, results=results)

        resultbook.save(outputpath)
    finally:
        if resultbook:
            resultbook.close()
