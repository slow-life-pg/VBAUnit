import json
import shutil
from collections import namedtuple
from pathlib import Path
from enum import StrEnum
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table
from pathutil.util import gettooldir


class ConfigPlace(StrEnum):
    LOCAL = "local"
    TOOL = "tool"
    OUTSIDE = "outside"


ScenarioElement = namedtuple("ScenarioElement", ["id", "subject", "module"])


def getlocalscenariofilenam() -> str:
    return "localfile.xlsx"


def createconfigfile(place: ConfigPlace) -> Path:
    configpath = __getactualpath(place=place)
    jsonobj = {"scenario": getlocalscenariofilenam()}
    with open(str(configpath), mode="wt", encoding="utf-8") as fp:
        json.dump(obj=jsonobj, fp=fp)
    return configpath


def deleteconfigfile(place: ConfigPlace) -> None:
    configpath = __getactualpath(place=place)
    if configpath.exists():
        configpath.unlink()
    if place == ConfigPlace.OUTSIDE:
        configdir = configpath.parent
        shutil.rmtree(configdir, ignore_errors=True)


def createscenariofile(scenarioobj: list[tuple[str, list[ScenarioElement]]]) -> Path:
    testingtempdir = Path("testing").resolve()
    if not testingtempdir.exists():
        testingtempdir.mkdir()
    scenariopath = testingtempdir.joinpath("scenario.xlsx")

    book = Workbook()
    defaultsheets = book.sheetnames

    tableid = 1
    for group in scenarioobj:
        print(f"create sheet: {group[0]}")
        sheet: Worksheet = book.create_sheet(group[0])
        sheet.cell(2, 2).value = "テストID"
        sheet.cell(2, 3).value = "説明"
        sheet.cell(2, 4).value = "モジュール"
        row = 3
        for testcase in group[1]:
            print(f"case: {testcase}")
            sheet.cell(row, 2).value = testcase.id
            sheet.cell(row, 3).value = testcase.subject
            sheet.cell(row, 4).value = testcase.module
            row += 1
        table = Table(id=tableid, ref=sheet.dimensions, displayName=f"Table{tableid}")
        sheet.tables.add(table=table)
        tableid += 1

    if len(book.worksheets) > len(defaultsheets):
        for title in defaultsheets:
            del book[title]
    print(f"sheets: {book.worksheets}")
    book.save(str(scenariopath))
    book.close()

    return scenariopath


def deletefile(filepath: Path) -> None:
    if filepath.exists():
        filepath.unlink()


def deletetestfiles() -> None:
    testingtempdir = Path("testing").resolve()
    if testingtempdir.exists():
        shutil.rmtree(testingtempdir)


def __getactualpath(place: ConfigPlace) -> Path:
    if place == ConfigPlace.LOCAL:
        return __getlocalconfigpath()
    elif place == ConfigPlace.TOOL:
        return __gettoolconfigpath()
    elif place == ConfigPlace.OUTSIDE:
        return __getlocalconfigpath()
    else:
        return Path.cwd()


def __getlocalconfigpath() -> Path:
    localdir = Path.cwd()
    localconfigpath = localdir.joinpath("config.json")
    return localconfigpath


def __gettoolconfigpath() -> Path:
    tooldir = gettooldir()
    toolconfigpath = tooldir.joinpath("config.json")
    return toolconfigpath
